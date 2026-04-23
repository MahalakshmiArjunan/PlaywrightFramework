import openpyxl

class ExcelUtility:
    def getRowCount(self, fileName, sheetName):
        workbook = openpyxl.load_workbook(fileName)
        sheet = workbook[sheetName]
        return sheet.max_row

    def getColumnCount(self, fileName, sheetName):
        workbook = openpyxl.load_workbook(fileName)
        sheet = workbook[sheetName]
        return sheet.max_column

    def getCellData(self, fileName, sheetName, rowNumber, columnNumber):
        workbook = openpyxl.load_workbook(fileName)
        sheet = workbook[sheetName]
        return sheet.cell(row=rowNumber, column=columnNumber).value

    def setCellData(self, fileName, sheetName, rowNumber, columnNumber, inputData):
        workbook = openpyxl.load_workbook(fileName)
        sheet = workbook[sheetName]
        sheet.cell(row=rowNumber, column=columnNumber).value = inputData
        workbook.save(fileName)
